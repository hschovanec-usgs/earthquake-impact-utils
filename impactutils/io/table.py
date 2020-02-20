# stdlib imports
import json
import re
import string
import time

# third party imports
import pandas as pd
import numpy as np
from lxml import etree
from openpyxl import load_workbook, utils

REQUIRED_COLUMNS = ['STATION', 'LAT', 'LON', 'NETID']
CHANNEL_GROUPS = [['[A-Z]{2}E', '[A-Z]{2}N', '[A-Z]{2}Z'],
                  ['[A-Z]{2}1', '[A-Z]{2}2', '[A-Z]{2}Z'],
                  ['H1', 'H2', 'Z'],
                  ['UNK']]
CHANNEL_PATTERNS = ['^[H,B][H,L,N][E,N,Z,1,2,3]$',  # match standard seed names
                    '^H[1,2]$',  # match H1/H2
                    '^Z$']  # match Z
PGM_COLS = ['PGA', 'PGV', 'SA(0.3)', 'SA(1.0)', 'SA(3.0)']
OPTIONAL = ['NAME', 'DISTANCE', 'REFERENCE',
            'INTENSITY', 'SOURCE', 'LOC', 'INSTTYPE', 'ELEV',
            'NRESP', 'INTENSITY_STDDEV']
FLOATRE = "[-+]?[0-9]*\.?[0-9]+"


def _move(cellstr, nrows, ncols):
    """Internal method for adding rows/columns to cell coordinate.

    'A1' moved by 1 row, 1 column => 'B2'

    Args:
        cellstr (str): Cell coordinate (A1,B2)
        nrows (int): Number of rows to move (usually down)
        ncols (int): Number of columns to move (usually right)
    Returns:
        str: New cell coordinate.
    """
    # WARNING! This will only work up to column Z!
    # colidx is a string, rowidx is a number
    col_str_idx, rowidx = utils.cell.coordinate_from_string(cellstr)
    letters = string.ascii_uppercase
    try:
        colidx = letters.index(col_str_idx)
        newcolidx = colidx + ncols
        newrowidx = rowidx + nrows
        newcellstr = f'{letters[newcolidx]}{int(newrowidx):d}'
        return newcellstr
    except ValueError:
        raise ValueError(
            f'Could not add {int(ncols):d} columns to column {col_str_idx}.')


def read_excel(excelfile):
    """Read strong motion Excel spreadsheet, return a DataFrame.

    Args:
        excelfile (str): Path to valid Excel file.
    Returns:
        DataFrame: Multi-indexed dataframe as described below.
        str or None: Reference string or None.

     - "STATION" String containing UNIQUE identifying station information.
     - "LAT" Latitude where peak ground motion observations were made.
     - "LON" Longitude where peak ground motion observations were made.
     - "NETID" (usually) two letter code indicating the source network.

    Optional columns include:
     - "NAME" String describing area where peak ground motions were observed.
     - "SOURCE" String describing (usu. long form) source of peak ground
       motion data.
     - "DISTANCE" Distance from epicenter to station location, in units of km.
     - "LOC" Two character location code.
     - "INSTTYPE" Instrument type, str.
     - "ELEV" Station elevation, in meters.

    And then at least one of the following columns:
     - "INTENSITY" MMI value (1-10).

        AND/OR
      a grouped set of per-channel peak ground motion columns, like this:

      -------------------------------------------------------------------------------
      |         H1              |           H2            |             Z           |
      -------------------------------------------------------------------------------
      |pga|pgv|psa03|psa10|psa30|pga|pgv|psa03|psa10|psa30|pga|pgv|psa03|psa10|psa30|
      -------------------------------------------------------------------------------

      The peak ground motion columns can be any of the following:
      - "PGA" Peak ground acceleration in units of %g.
      - "PGV" Peak ground velocity in units of cm/sec.
      - "PSA03" Peak spectral acceleration at 0.3 seconds, in units of %g.
      - "PSA10" Peak spectral acceleration at 1.0 seconds, in units of %g.
      - "PSA30" Peak spectral acceleration at 3.0 seconds, in units of %g.

    Valid "channel" columns are {H1,H2,Z} or {XXN,XXE,XXZ}, where 'XX' is any
    two-letter combination, usually adhering to the following standard:
    http://www.fdsn.org/seed_manual/SEEDManual_V2.4_Appendix-A.pdf

    If the input data set provides no channel information, then the channel
    can be simply "UNK".

    """

    # figure out if data frame is multi-index or not
    wb = load_workbook(excelfile)
    ws = wb.active

    # figure out where the top left of the data begins
    topleft = 'A1'
    first_cell = 'A2'
    second_cell = 'A3'

    # figure out if there is a little reference section in this...
    reference = None
    skip_rows = None
    header = [0, 1]
    if ws[topleft].value.lower() != 'reference':
        raise KeyError('Reference cells are required in A1 and B1!')
    refcell = _move(topleft, 0, 1)
    reference = ws[refcell].value
    first_cell = _move(topleft, 1, 0)
    second_cell = _move(first_cell, 1, 0)
    skip_rows = [0]
    header = [1, 2]

    is_multi = True
    # if the first column of the second row is not empty,
    # then we do not have a multi-index.
    if ws[second_cell].value is not None:
        is_multi = False

    # read in dataframe, assuming that ground motions are grouped by channel
    if is_multi:
        try:
            # note - in versions of pandas prior to 0.24, index_col=None
            # has no effect here.  Hence the unsetting of the index later.
            df = pd.read_excel(excelfile, header=header, index_col=None)
            # if the name column is all blanks, it's filled with NaNs by
            # default, which causes problems later on.  Replace with
            # empty strings
            if 'NAME' in df.columns:
                df['NAME'] = df['NAME'].fillna('')
        except pd.errors.ParserError:
            raise IndexError('Input file has invalid empty first data row.')

        headers = df.columns.get_level_values(0).str.upper()
        subheaders = df.columns.get_level_values(1).str.upper()
        df.columns = pd.MultiIndex.from_arrays([headers, subheaders])
        top_headers = df.columns.levels[0]
    else:
        df = pd.read_excel(excelfile, skiprows=skip_rows, index_col=None)
        top_headers = df.columns

    # make sure basic columns are present
    if 'STATION' not in top_headers:
        df['STATION'] = df.index
        df = df.reset_index(drop=True)
        top_headers = df.columns.levels[0]
    if not set(REQUIRED_COLUMNS).issubset(set(top_headers)):
        fmt = f'Input Excel file must specify the following columns: {(str(REQUIRED_COLUMNS))}.'
        raise KeyError(fmt)

    # check if channel headers are valid
    channels = (set(top_headers) - set(REQUIRED_COLUMNS)) - set(OPTIONAL)
    valid = False
    if len(channels):
        for channel_group in CHANNEL_GROUPS:
            num_channels = 0
            for channel_pat in channel_group:
                cp = re.compile(channel_pat)
                if len(list(filter(cp.match, channels))):
                    num_channels += 1
            if num_channels == 1 and len(channels) == 1:
                valid = True
                break
            elif num_channels > 1:
                h1_pat = re.compile(channel_group[0])
                h2_pat = re.compile(channel_group[1])
                has_h1 = len(list(filter(h1_pat.match, channels))) > 0
                has_h2 = len(list(filter(h2_pat.match, channels))) > 0
                if has_h1 or has_h2:
                    valid = True
                    break
    else:
        valid = True
    if not valid:
        raise KeyError(
            f'{str(sorted(list(channels)))} is not a valid channel grouping')

    # make sure the empty cells are all nans or floats
    found = False
    if 'INTENSITY' in top_headers:
        found = True
    empty_cell = re.compile(r'\s+')
    for channel in channels:
        channel_df = df[channel].copy()
        for column in PGM_COLS:
            if column in channel_df:
                found = True
                channel_df[column] = channel_df[column].replace(
                    empty_cell, np.nan)
                channel_df[column] = channel_df[column].astype(float)
        df[channel] = channel_df

    if not found:
        intensity_col = str(PGM_COLS + ['intensity'])
        fmt = (f'File must contain at least one of the following '
               f'data columns: {intensity_col}')
        raise KeyError(fmt)

    return (df, reference)


def _get_channels(columns):
    channels = []
    for column in columns:
        for cmatch in CHANNEL_PATTERNS:
            if re.search(cmatch, column) is not None:
                channels.append(column)
                break
    return channels


def dataframe_to_json(df, jsonfile, earthquake_location='--'):
    """Write a dataframe to json format.

    Args:
        df (DataFrame): Pandas dataframe, as described in read_excel.
        jsonfile (str): Path to file where XML file should be written.

    Notes:
        This method accepts either a dataframe from read_excel, or
        one with this structure:
         - STATION: Station code (REQUIRED)
         - IMC: Component (HHE,HHN, GREATER_OF_TWO, ROTD50, etc.) (REQUIRED)
         - IMT: Intensity measure type (pga,pgv, etc.) (REQUIRED)
         - VALUE: IMT value. (REQUIRED)
         - LAT: Station latitude. (REQUIRED)
         - LON: Station longitude. (REQUIRED)
         - NETID: Station contributing network. (REQUIRED)
         - FLAG: String quality flag, meaningful to contributing networks,
                 but ShakeMap ignores any station with a non-zero value. (REQUIRED)
         - ELEV: Elevation of station (m). (OPTIONAL)
         - NAME: String describing station. (OPTIONAL)
         - DISTANCE: Distance (km) from station to origin. (OPTIONAL)
         - LOC: Description of location (i.e., "5 km south of Wellington")
                (OPTIONAL)
         - INSTTYPE: Instrument type (FBA, etc.) (OPTIONAL)
         - INTENSITY: MMI intensity. (OPTIONAL)
         - NRESP: Number of responses for aggregated intensity. (OPTIONAL)
         - INTENSITY_STDDEV: Uncertainty for this intensity. (OPTIONAL)
    """
    features = []

    if hasattr(df.columns, 'levels'):
        top_headers = set(df.columns.levels[0])
        required = set(REQUIRED_COLUMNS)
        optional = set(OPTIONAL)
        channel_candidates = (top_headers - required) - optional
        channels = _get_channels(channel_candidates)
    else:
        channels = []

    processed_stations = []
    for _, row in df.iterrows():
        tmprow = row.copy()
        if isinstance(tmprow.index, pd.core.indexes.multi.MultiIndex):
            tmprow.index = tmprow.index.droplevel(1)

        # assign required columns
        stationcode = str(tmprow['STATION']).strip()

        netid = tmprow['NETID'].strip()
        if not stationcode.startswith(netid):
            stationcode = f'{netid}.{stationcode}'

        # if this is a dataframe created by shakemap,
        # there will be multiple rows per station.
        # below we process all those rows at once,
        # so we need this bookkeeping to know that
        # we've already dealt with this station
        if stationcode in processed_stations:
            continue

        station = {}

        ## Geometry
        geom = {}
        geom['type'] = 'Point'
        lat = f"{tmprow['LAT']:.4f}"
        lon = f"{tmprow['LON']:.4f}"
        elev = f"{tmprow['ELEV']:.1f}"
        geom['coordinates'] = [
            lat, lon, elev
        ]
        station['geometry'] = geom

        ## Properties
        props = {}
        props['code'] = stationcode.split('.')[-1]
        # standard column headers
        if 'NAME' in tmprow:
            props['name'] = tmprow['NAME'].strip()
        if 'NETID' in tmprow:
            props['network'] = tmprow['NETID'].strip()
        if 'DISTANCE' in tmprow:
            props['distance'] = f"{tmprow['DISTANCE']:.1f}"
        if 'INTENSITY' in tmprow:
            props['intensity'] = f"{tmprow['INTENSITY']:.1f}"
        if 'NRESP' in tmprow:
            props['nresp'] = f"{int(tmprow['NRESP']):d}"
        if 'INTENSITY_STDDEV' in tmprow:
            props['intensity_stddev'] = f"{tmprow['INTENSITY_STDDEV']:.2f}"
        if 'SOURCE' in tmprow:
            props['source'] = tmprow['SOURCE'].strip()
        if 'LOC' in tmprow:
            props['location'] = tmprow['LOC'].strip()
        else:
            props['location'] = '--'
        if 'INSTTYPE' in tmprow:
            props['type'] = tmprow['INSTTYPE'].strip()
        # new format properties
        if 'PROVIDER' in tmprow:
            props['provider'] = tmprow['PROVIDER'].strip()
        if 'INSTRUMENT' in tmprow:
            props['instrument'] = tmprow['INSTRUMENT'].strip()
        if 'SERIAL' in tmprow:
            props['serial'] = tmprow['SERIAL'].strip()
        else:
            props['serial'] = 'None'
        if 'PERIOD' in tmprow:
            props['period'] = tmprow['PERIOD'].strip()
        if 'DAMPING' in tmprow:
            props['damping'] = tmprow['DAMPING'].strip()
        if 'SENSITIVITY' in tmprow:
            props['sensitivity'] = tmprow['SENSITIVITY'].strip()
        if 'SOURCE_FORMAT' in tmprow:
            props['source_format'] = tmprow['SOURCE_FORMAT'].strip()
        if 'STRUCTURE' in tmprow:
            props['structure'] = tmprow['STRUCTURE'].strip()
        station['properties'] = props


        station['channels'] = {}
        # sort channels by N,E,Z or H1,H2,Z
        channels = sorted(list(channels))

        for channel in channels:
            if channel not in station['channels']:
                station['channels'][channel.upper()] = {}
                station['channels'][channel.upper()]['amplitudes'] = []

            for pgm in ['pga', 'pgv', 'psa03', 'psa10', 'psa30']:
                newpgm = _translate_imt(pgm, row[channel].index)
                c1 = newpgm not in row[channel]
                c2 = False
                if not c1:
                    c2 = np.isnan(row[channel][newpgm])
                if c1 or c2:
                    continue
                # make an element with the old style name
                p = {'name': pgm, 'flag':0, 'value': f'{row[channel][newpgm]:.4f}'}
                station['channels'][channel.upper()]['amplitudes'] += [p]
        features += [station]
        processed_stations.append(stationcode)
    import pprint
    geojson = {
        'type': "FeatureCollection",
        'software': {
            'name': "impactutils",
            'version': "1.0.0"
        },
        'process_time': int(time.time()),
        'features': features
    }
    with open('test.json', 'w') as f:
        json.dump(geojson, f, indent=4)

def dataframe_to_xml(df, xmlfile, reference=None):
    """Write a dataframe to ShakeMap XML format.

    This method accepts either a dataframe from read_excel, or
    one with this structure:
     - STATION: Station code (REQUIRED)
     - CHANNEL: Channel (HHE,HHN, etc.) (REQUIRED)
     - IMT: Intensity measure type (pga,pgv, etc.) (REQUIRED)
     - VALUE: IMT value. (REQUIRED)
     - LAT: Station latitude. (REQUIRED)
     - LON: Station longitude. (REQUIRED)
     - NETID: Station contributing network. (REQUIRED)
     - FLAG: String quality flag, meaningful to contributing networks,
             but ShakeMap ignores any station with a non-zero value. (REQUIRED)
     - ELEV: Elevation of station (m). (OPTIONAL)
     - NAME: String describing station. (OPTIONAL)
     - DISTANCE: Distance (km) from station to origin. (OPTIONAL)
     - LOC: Description of location (i.e., "5 km south of Wellington")
            (OPTIONAL)
     - INSTTYPE: Instrument type (FBA, etc.) (OPTIONAL)
     - INTENSITY: MMI intensity. (OPTIONAL)
     - NRESP: Number of responses for aggregated intensity. (OPTIONAL)
     - INTENSITY_STDDEV: Uncertainty for this intensity. (OPTIONAL)

    Args:
        df (DataFrame): Pandas dataframe, as described in read_excel.
        xmlfile (str): Path to file where XML file should be written.
    """
    if hasattr(df.columns, 'levels'):
        top_headers = set(df.columns.levels[0])
        required = set(REQUIRED_COLUMNS)
        optional = set(OPTIONAL)
        channel_candidates = (top_headers - required) - optional
        channels = _get_channels(channel_candidates)
    else:
        channels = []
    root = etree.Element('shakemap-data', code_version="3.5", map_version="3")

    create_time = int(time.time())
    stationlist = etree.SubElement(
        root, 'stationlist', created=f'{int(create_time):d}')
    if reference is not None:
        stationlist.attrib['reference'] = reference

    processed_stations = []

    for _, row in df.iterrows():
        tmprow = row.copy()
        if isinstance(tmprow.index, pd.core.indexes.multi.MultiIndex):
            tmprow.index = tmprow.index.droplevel(1)

        # assign required columns
        stationcode = str(tmprow['STATION']).strip()

        netid = tmprow['NETID'].strip()
        if not stationcode.startswith(netid):
            stationcode = f'{netid}.{stationcode}'

        # if this is a dataframe created by shakemap,
        # there will be multiple rows per station.
        # below we process all those rows at once,
        # so we need this bookkeeping to know that
        # we've already dealt with this station
        if stationcode in processed_stations:
            continue

        station = etree.SubElement(stationlist, 'station')

        station.attrib['code'] = stationcode
        station.attrib['lat'] = f"{tmprow['LAT']:.4f}"
        station.attrib['lon'] = f"{tmprow['LON']:.4f}"

        # assign optional columns
        if 'NAME' in tmprow:
            station.attrib['name'] = tmprow['NAME'].strip()
        if 'NETID' in tmprow:
            station.attrib['netid'] = tmprow['NETID'].strip()
        if 'DISTANCE' in tmprow:
            station.attrib['dist'] = f"{tmprow['DISTANCE']:.1f}"
        if 'INTENSITY' in tmprow:
            station.attrib['intensity'] = f"{tmprow['INTENSITY']:.1f}"
        if 'NRESP' in tmprow:
            station.attrib['nresp'] = f"{int(tmprow['NRESP']):d}"
        if 'INTENSITY_STDDEV' in tmprow:
            station.attrib['intensity_stddev'] = f"{tmprow['INTENSITY_STDDEV']:.2f}"
        if 'SOURCE' in tmprow:
            station.attrib['source'] = tmprow['SOURCE'].strip()
        if 'LOC' in tmprow:
            station.attrib['loc'] = tmprow['LOC'].strip()
        if 'INSTTYPE' in tmprow:
            station.attrib['insttype'] = tmprow['INSTTYPE'].strip()
        if 'ELEV' in tmprow:
            station.attrib['elev'] = f"{tmprow['ELEV']:.1f}"

        if 'imt' not in tmprow.index:
            # sort channels by N,E,Z or H1,H2,Z
            channels = sorted(list(channels))

            for channel in channels:
                component = etree.SubElement(station, 'comp')
                component.attrib['name'] = channel.upper()

                # figure out if channel is horizontal or vertical
                if channel[-1] in ['1', '2', 'E', 'N']:
                    component.attrib['orientation'] = 'h'
                else:
                    component.attrib['orientation'] = 'z'

                # create sub elements out of any of the PGMs
                # this is extra confusing because we're trying to
                # transition from psa03 style to SA(0.3) style.
                # station xml format only accepts the former, but we're
                # supporting the latter as input, and the format as output.

                # loop over desired output fields
                for pgm in ['pga', 'pgv', 'psa03', 'psa10', 'psa30']:
                    newpgm = _translate_imt(pgm, row[channel].index)
                    c1 = newpgm not in row[channel]
                    c2 = False
                    if not c1:
                        c2 = np.isnan(row[channel][newpgm])
                    if c1 or c2:
                        continue
                    # make an element with the old style name
                    pgm_el = etree.SubElement(component, pgm)
                    pgm_el.attrib['flag'] = '0'
                    pgm_el.attrib['value'] = f'{row[channel][newpgm]:.4f}'
            processed_stations.append(stationcode)
        else:
            # this file was created by a process that has imt/value columns
            # search the dataframe for all rows with this same station code
            scode = tmprow['STATION']
            station_rows = df[df['STATION'] == scode]

            # now we need to find all of the channels
            channels = station_rows['channel'].unique()
            for channel in channels:
                channel_rows = station_rows[station_rows['channel'] == channel]
                component = etree.SubElement(station, 'comp')
                component.attrib['name'] = channel.upper()
                for _, channel_row in channel_rows.iterrows():
                    pgm = channel_row['imt']
                    value = channel_row['value']

                    pgm_el = etree.SubElement(component, pgm)
                    pgm_el.attrib['value'] = f'{value:.4f}'
                    pgm_el.attrib['flag'] = str(channel_row['flag'])

            processed_stations.append(stationcode)

    tree = etree.ElementTree(root)
    tree.write(xmlfile, pretty_print=True)


def _translate_imt(oldimt, imtlist):
    # translate from psa03 to sa(0.3)
    if oldimt.upper() in ['PGA', 'PGV']:
        newimt = oldimt.upper()
    else:
        match = re.search(r'\d+', oldimt)
        if match is not None:
            period = float(match.group()) / 10
            for imt in imtlist:
                if not imt.startswith('SA'):
                    continue
                try:
                    imt_period = float(re.search(FLOATRE, imt).group())
                except Exception:
                    continue
                if imt_period == period:
                    newimt = imt
                    break
        else:
            newimt = ''
    return newimt
